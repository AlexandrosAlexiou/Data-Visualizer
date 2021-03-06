#!/usr/bin/env python3
import openpyxl as xl
import csv
import sys

# Global variables
xl__files = {
    'Austria': 'Data/API_AUT_DS2_en_excel_v2_827542.xlsx',
    'Denmark': 'Data/API_DNK_DS2_en_excel_v2_825118.xlsx',
    'Greece': 'Data/API_GRC_DS2_en_excel_v2_827586.xlsx',
    'Italy': 'Data/API_ITA_DS2_en_excel_v2_827758.xlsx',
    'Japan': 'Data/API_JPN_DS2_en_excel_v2_822387.xlsx',
    'Korea': 'Data/API_KOR_DS2_en_excel_v2_823838.xlsx',
    'Switzerland': 'Data/API_CHE_DS2_en_excel_v2_1011017.xlsx',
    'Sweden': 'Data/API_SWE_DS2_en_excel_v2_825541.xlsx',
    'Turkey': 'Data/API_TUR_DS2_en_excel_v2_821151.xlsx',
    'United States': 'Data/API_USA_DS2_en_excel_v2_996440.xlsx'
}
five_year_period   = ['1960-1964', '1965-1969', '1970-1974', '1975-1979', '1980-1984', '1985-1989',
                    '1990-1994', '1995-1999', '2000-2004', '2005-2009', '2010-2014', '2015-2019']

ten_year_period    = ['1960-1969', '1970-1979', '1980-1989', '1990-1999', '2000-2009', '2010-2019']

twenty_year_period = ['1960-1979', '1980-1999', '2000-2019']

row = 1

# Data config function definitions


def create_measurements_xl_file():
    global row

    filepath = "Exported_data/measurements.xlsx"
    wb = xl.Workbook()
    wb.save(filepath)

    # opening the destination excel file
    output_xl_file = 'Exported_data/measurements.xlsx'
    wb2 = xl.load_workbook(output_xl_file)
    ws2 = wb2.active

    for country in xl__files:
        wb1 = xl.load_workbook(xl__files[country])
        ws1 = wb1.worksheets[0]

        # calculate total number of rows and columns in source excel file
        max_rows = ws1.max_row

        for rows in range(5, max_rows + 1):
            # reading cell value from source excel file
            c1 = ws1.cell(row=rows, column=2)
            c2 = ws1.cell(row=rows, column=4)
            if c2.value == 'SP.DYN.LE00.MA.IN' or c2.value == 'TX.VAL.MRCH.WL.CD' or \
                    c2.value == 'SP.POP.2024.FE.5Y' or c2.value == 'SP.DYN.TO65.MA.ZS' or \
                    c2.value == 'NY.GSR.NFCY.CD' or c2.value == 'EN.ATM.CO2E.SF.ZS' or \
                    c2.value == 'EG.ELC.FOSL.ZS' or c2.value == 'MS.MIL.XPND.CD' or \
                    c2.value == 'NV.MNF.FBTO.ZS.UN' or c2.value == 'EN.URB.LCTY.UR.ZS' or \
                    c2.value == 'BX.GSR.GNFS.CD' or c2.value == 'BM.GSR.GNFS.CD':
                for columns in range(5, 65):
                  # writing the read value to destination excel file
                  ws2.cell(row, 1).value = c1.value
                  ws2.cell(row, 2).value = str(c2.value)

                  # reading year values from source excel file
                  c3 = ws1.cell(4, columns)

                  # writing years value to destination excel file
                  ws2.cell(row, 3).value = c3.value

                  # reading measurement values from source excel file
                  c4 = ws1.cell(rows, columns)

                  if ws1.cell(rows, columns).value is not None:
                    # writing measurement values to destination excel file
                    ws2.cell(row, 4).value = c4.value
                  else:
                    # there is no measurement so the value is 0
                     ws2.cell(row, 4).value = 0

                  # destination file new row
                  row += 1
    # saving the destination excel file
    wb2.save(str(output_xl_file))


def create_indicators_xl_file():
    wb1 = xl.load_workbook('Data/API_AUT_DS2_en_excel_v2_827542.xlsx')
    ws1 = wb1.worksheets[0]

    # creating xlsx file
    filepath = "Exported_data/indicators.xlsx"
    wb = xl.Workbook()
    wb.save(filepath)

    # opening the destination excel file
    output_xl_file = 'Exported_data/indicators.xlsx'
    wb2 = xl.load_workbook(output_xl_file)
    ws2 = wb2.active

    # calculate total number of rows and columns in source excel file
    max_rows = ws1.max_row

    # Begin to write from the second row
    row_number = 1
    for rows in range(5, max_rows + 1):
        # reading cell value from source excel file
        c1 = ws1.cell(row=rows, column=3)
        c2 = ws1.cell(row=rows, column=4)
        if c2.value == 'SP.DYN.LE00.MA.IN' or c2.value == 'TX.VAL.MRCH.WL.CD' or \
                c2.value == 'SP.POP.2024.FE.5Y' or c2.value == 'SP.DYN.TO65.MA.ZS' or \
                c2.value == 'NY.GSR.NFCY.CD' or c2.value == 'EN.ATM.CO2E.SF.ZS' or \
                c2.value == 'EG.ELC.FOSL.ZS' or c2.value == 'MS.MIL.XPND.CD' or \
                c2.value == 'NV.MNF.FBTO.ZS.UN' or c2.value == 'EN.URB.LCTY.UR.ZS' or \
                c2.value == 'BX.GSR.GNFS.CD' or c2.value == 'BM.GSR.GNFS.CD':
            ws2.cell(row_number, 1).value = c2.value

            ws2.cell(row_number, 2).value = str(c1.value)

            # destination file new row
            row_number += 1
    # saving the destination excel file
    wb2.save(str(output_xl_file))


def create_countries_xl_file():
    # creating xlsx file
    filepath = "Exported_data/countries.xlsx"
    wb = xl.Workbook()
    wb.save(filepath)
    # opening the destination excel file
    output_xl_file = 'Exported_data/countries.xlsx'
    wb2 = xl.load_workbook(output_xl_file)
    ws2 = wb2.active
    row_number = 1
    for country in xl__files:
        wb1 = xl.load_workbook(xl__files[country])
        ws1 = wb1.worksheets[0]
        c1 = ws1.cell(5, 1)
        c2 = ws1.cell(5, 2)
        ws2.cell(row_number, 1).value = c2.value
        ws2.cell(row_number, 2).value = str(c1.value)
        row_number += 1
    # saving the destination excel file
    wb2.save(str(output_xl_file))


def create_years_xl_file():

    global five_year_period, ten_year_period, twenty_year_period

    wb1 = xl.load_workbook('Data/API_AUT_DS2_en_excel_v2_827542.xlsx')
    ws1 = wb1.worksheets[0]
    # creating xlsx file
    filepath = "Exported_data/years.xlsx"
    wb = xl.Workbook()
    wb.save(filepath)
    # opening the destination excel file
    output_xl_file = 'Exported_data/years.xlsx'
    wb2 = xl.load_workbook(output_xl_file)
    ws2 = wb2.active

    # Begin to write from the second row
    row_number = 1
    for columns in range(5, 65):
        c1 = ws1.cell(4, columns)
        ws2.cell(row_number, 1).value = c1.value
        row_number += 1
    row_number = 1
    for year in range(0, 12):
        for i in range(1, 6):
            ws2.cell(row_number, 2).value = five_year_period[year]
            row_number += 1
    row_number = 1
    for year in range(0, 6):
        for i in range(1, 11):
            ws2.cell(row_number, 3).value = ten_year_period[year]
            row_number += 1
    row_number = 1
    for year in range(0, 3):
        for i in range(1, 21):
            ws2.cell(row_number, 4).value = twenty_year_period[year]
            row_number += 1

    wb2.save(str(output_xl_file))


def convert_to_csv(filename):
    wb = xl.load_workbook('Exported_data/'+filename+'.xlsx')
    sh = wb.active
    with open('Exported_data/csv/'+filename+'.csv', 'w') as f:
        c = csv.writer(f, quoting=csv.QUOTE_ALL)
        for r in sh.rows:
            c.writerow([cell.value for cell in r])
    f.close()
    with open('Exported_data/csv/'+filename+'.csv', 'rb+') as f:
        content = f.read()
        f.seek(0)
        f.write(content.replace(b'\r', b''))
        f.truncate()


# Driver program
def main():


    create_years_xl_file()
    print("Created years.xlsx file. ")
    create_measurements_xl_file()
    print("Created measurements.xlsx file. ")
    create_countries_xl_file()
    print("Created countries.xlsx file. ")
    create_indicators_xl_file()
    print("Created indicators.xlsx file. ")
    convert_to_csv('countries')
    convert_to_csv('indicators')
    convert_to_csv('measurements')
    convert_to_csv('years')
    print("Converted all files to csv.")


if __name__ == '__main__':
    main()
